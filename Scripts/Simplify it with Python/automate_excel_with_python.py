#!/usr/bin/env python
# coding: utf-8

# In[285]:


import pandas as pd
import numpy as np
from openpyxl import load_workbook, workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side


# In[272]:


# Import the data
mart = pd.read_excel(r"D:\Learnerea\Tables\mart_train_sample.xlsx")
mart.head()


# In[273]:


# Summarizing the data
summary = mart.pivot_table(values=['Sales'],index=['Outlet_Year'],columns=['Outlet_Location_Type'])
summary


# In[274]:


# Export summary to excel
summary.to_excel(r"D:\Learnerea\Tables\mart_summary.xlsx")


# In[275]:


# Load excel and creat an workbook object and then the worksheet object
wb = load_workbook(r"D:\Learnerea\Tables\mart_summary.xlsx")
ws = wb['Sheet1']


# In[276]:


# Unmerge the merged cells
ws.unmerge_cells("b1:d1")


# In[277]:


# Delete rows
ws.delete_rows(3)
ws.delete_rows(1)


# In[278]:


# Testing
ws['g1'] = 'my name is learnerea'
ws['g1'] = ""


# In[289]:


# Style templates
header_fill = PatternFill(patternType='solid',fgColor='d66058')
heading_fill = PatternFill(patternType='solid',fgColor='62b7f0')

hader_font = Font(bold=True, color='f5f8fa')
heading_font = Font(name='Lemon', bold=True, color='f5f8fa', size=15)

border_style = Border(left=Side(style='thick'),right=Side(style='thick'),top=Side(style='thick'),bottom=Side(style='thick'))


# In[280]:


# Looping through range of chells and applying the formats
for row in range(1,2):
    for col in range(1,5):
        col_name = get_column_letter(col)
        ws[col_name+str(row)].fill = header_fill
        ws[col_name+str(row)].font = hader_font
        
for row in range(1,12):
    for col in range(1,2):
        col_name = get_column_letter(col)
        ws[col_name+str(row)].fill = header_fill
        ws[col_name+str(row)].font = hader_font


# In[281]:


# Insert rows and columns
ws.insert_cols(1,5)
ws.insert_rows(1,7)


# In[282]:


# merging celss
ws.merge_cells("f6:i7")


# In[283]:


# apply format to heading
ws['f6'] = "MART SALES DATA"
ws['f6'].alignment = Alignment(horizontal='center',vertical='center')
ws['f6'].fill = heading_fill
ws['f6'].font = heading_font


# In[290]:


# Set Border
for row in range(6,19):
    for col in range(6,10):
        col_name = get_column_letter(col)
        ws[col_name+str(row)].border =    border_style


# In[292]:


# Freeze panes
freez_cell = ws['g9']
ws.freeze_panes=freez_cell


# In[293]:


# Export or save the data
wb.save(r"D:\Learnerea\Tables\mart_summary2.xlsx")


# In[ ]:




